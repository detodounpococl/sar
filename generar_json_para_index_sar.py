import json
import os
from collections import defaultdict
from openpyxl import load_workbook

CARPETA = r"C:\Users\14538348-k\Desktop\DASHBOARD (html)\Nuevos\SAR"
ARCHIVO_SALIDA = "sar_dashboard.json"

MESES = [
    "ENERO","FEBRERO","MARZO","ABRIL","MAYO","JUNIO",
    "JULIO","AGOSTO","SEPTIEMBRE","OCTUBRE","NOVIEMBRE","DICIEMBRE"
]

MAPA_MESES = {
    "ENERO": 1, "FEBRERO": 2, "MARZO": 3, "ABRIL": 4,
    "MAYO": 5, "JUNIO": 6, "JULIO": 7, "AGOSTO": 8,
    "SEPTIEMBRE": 9, "SETIEMBRE": 9, "OCTUBRE": 10,
    "NOVIEMBRE": 11, "DICIEMBRE": 12
}

FILA_TOTAL_INSCRITOS = 3
FILA_HORA_INICIO = 4
FILA_HORA_FIN = 27
COL_HORA = 1
COL_DIA_INICIO = 2
COL_DIA_FIN = 32

def clean_text(v):
    return "" if v is None else str(v).strip()

def safe_int(v):
    if v in (None, ""):
        return 0
    try:
        return int(float(str(v).replace(".", "").replace(",", ".")))
    except:
        return 0

def year_from_filename(filename):
    digits = "".join(ch for ch in os.path.basename(filename) if ch.isdigit())
    if len(digits) >= 4:
        return int(digits[:4])
    raise ValueError(f"No pude detectar el año en: {filename}")

def month_name(sheet_name):
    s = clean_text(sheet_name).upper()
    for m in MESES:
        if m in s:
            return m
    return None

def month_number(sheet_name):
    s = clean_text(sheet_name).upper()
    for m, n in MAPA_MESES.items():
        if m in s:
            return n
    return None

def is_month_sheet(name):
    s = clean_text(name).upper()
    if "MAPEO" in s or "CONSOLIDADO" in s or "RESUMEN" in s:
        return False
    return month_name(name) is not None

def hour_label(raw_hour):
    txt = clean_text(raw_hour)
    if not txt:
        return ""
    if ":" in txt:
        txt = txt.split(":")[0]
    try:
        return str(int(float(txt)))
    except:
        return txt

def main():
    salida = []

    for archivo in os.listdir(CARPETA):
        if not archivo.lower().endswith((".xlsm", ".xlsx")):
            continue

        ruta = os.path.join(CARPETA, archivo)
        anio = year_from_filename(archivo)

        print(f"Procesando {archivo}...")

        wb = load_workbook(ruta, data_only=True, read_only=True)

        for ws_name in wb.sheetnames:
            if not is_month_sheet(ws_name):
                continue

            ws = wb[ws_name]
            mes = month_name(ws_name)
            mes_num = month_number(ws_name)

            por_hora = {str(h): 0 for h in [8,9,10,11,12,13,14,15,16,17,18,19,20,21,22,23,0,1,2,3,4,5,6,7]}
            por_dia = {str(d): 0 for d in range(1, 32)}
            total_inscritos = 0

            for col in range(COL_DIA_INICIO, COL_DIA_FIN + 1):
                dia = col - 1
                total_dia = safe_int(ws.cell(FILA_TOTAL_INSCRITOS, col).value)
                total_inscritos += total_dia
                por_dia[str(dia)] = total_dia

            for fila in range(FILA_HORA_INICIO, FILA_HORA_FIN + 1):
                hora = hour_label(ws.cell(fila, COL_HORA).value)
                if hora == "":
                    continue

                total_hora = 0
                for col in range(COL_DIA_INICIO, COL_DIA_FIN + 1):
                    total_hora += safe_int(ws.cell(fila, col).value)

                por_hora[hora] = total_hora

            salida.append({
                "anio": anio,
                "mes": mes,
                "mes_num": mes_num,
                "total_inscritos": total_inscritos,
                "por_hora": por_hora,
                "por_dia": por_dia
            })

        wb.close()

    salida.sort(key=lambda x: (x["anio"], x["mes_num"]))

    with open(os.path.join(CARPETA, ARCHIVO_SALIDA), "w", encoding="utf-8") as f:
        json.dump(salida, f, ensure_ascii=False, indent=2)

    print("JSON completo generado correctamente")

if __name__ == "__main__":
    main()